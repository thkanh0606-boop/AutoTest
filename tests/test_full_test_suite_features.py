from pathlib import Path
from unittest.mock import MagicMock, patch

from openpyxl import load_workbook
import pytest

from core.suite_loader import dashboard_regression_cases, demo_7_module_cases
from core.pcm_suite_contract import pcm_suite_cases
from core.test_result_repository import TestResultRepository
from runners.pcm_suite_runner import SUPPORTED_SCENARIOS, run_pcm_scenario
from runners.suite_runner import SuiteWorker, _run_route_smoke
from services.suite_report_export import export_excel, export_html


TEST_DB = Path("data/test_suite_feature_test.sqlite3")


@pytest.fixture(autouse=True)
def cleanup_generated_reports():
    yield
    for path in (
        TEST_DB,
        Path("data/test_suite_report.html"),
        Path("data/test_suite_report.xlsx"),
    ):
        if path.exists():
            path.unlink()


def _fresh_repository():
    if TEST_DB.exists():
        TEST_DB.unlink()
    return TestResultRepository(str(TEST_DB))


def test_builtin_suites_cover_seven_modules_and_dashboard_contracts():
    demo = demo_7_module_cases()
    regression = dashboard_regression_cases()

    assert len(demo) == 7
    assert len({case["page_key"] for case in demo}) == 7
    assert all(case["action_type"] == "route_smoke" and case["executable"] for case in demo)
    assert regression
    assert all(case["locator_type"] and case["locator_value"] for case in regression)


def test_pcm_suite_has_all_31_automated_contracts():
    cases = pcm_suite_cases()

    assert [case["tc_id"] for case in cases] == [f"TC{number:02d}" for number in range(1, 32)]
    assert len({case["scenario_key"] for case in cases}) == 31
    assert {case["scenario_key"] for case in cases} == SUPPORTED_SCENARIOS
    assert all(case["action_type"] == "pcm_scenario" for case in cases)
    assert all(case["executable"] and case["expected"] for case in cases)


def test_unknown_pcm_scenario_is_error_without_false_pass():
    payload = run_pcm_scenario(
        {
            "tc_id": "TCXX",
            "scenario_key": "not_implemented",
            "module": "General",
        }
    )

    assert payload["status"] == "ERROR"
    assert "chưa được ánh xạ" in payload["error_message"]


def test_suite_worker_dispatches_pcm_contract_to_pcm_runner():
    case = pcm_suite_cases()[0]
    expected_payload = {
        "case_id": "TC01", "title": case["title"], "module": "Đăng nhập",
        "page_key": "plt_login", "expected": case["expected"], "actual": "/dashboard",
        "status": "PASSED", "message": "OK", "error_message": "",
        "screenshot_path": "", "log_text": "ASSERT", "started_at": "2026-08-14T00:00:00",
        "finished_at": "2026-08-14T00:00:01", "duration_ms": 1000,
    }
    worker = SuiteWorker([case])

    with patch("runners.pcm_suite_runner.run_pcm_scenario", return_value=expected_payload) as run:
        payload = worker._run_case(case)

    assert payload == expected_payload
    run.assert_called_once()


def test_unautomated_excel_case_is_skipped_not_passed():
    worker = SuiteWorker([])
    payload = worker._run_case(
        {
            "tc_id": "TC-EXCEL-01",
            "title": "Mô tả nghiệp vụ chưa có locator",
            "area": "Đặt xe",
            "page_key": "plt_booking",
            "expected": "Tạo booking thành công",
        }
    )

    assert payload["status"] == "SKIPPED"
    assert "không đánh PASS giả" in payload["message"]


def test_route_smoke_requires_expected_path_and_visible_dom():
    driver = MagicMock()
    driver.current_url = "https://courses.plt.pro.vn/login"
    driver.find_element.return_value.is_displayed.return_value = True
    case = demo_7_module_cases()[2]  # Booking expects /bookings.

    with (
        patch("core.driver_factory.DriverFactory.create_driver", return_value=driver),
        patch("runners.suite_runner.time.sleep"),
        patch("runners.suite_runner.capture_screenshot", return_value="booking.png"),
    ):
        payload = _run_route_smoke(case)

    assert payload["status"] == "FAILED"
    assert "actual /login" in payload["message"]
    assert payload["screenshot_path"] == "booking.png"
    driver.quit.assert_called_once()


def test_repository_groups_results_by_one_suite_run():
    repository = _fresh_repository()
    cases = demo_7_module_cases()[:2]
    suite_id = repository.save_suite_definition(
        "Demo unit", cases, suite_key="unit:demo"
    )
    run_id = repository.start_suite_run(suite_id, "Demo unit", "Selected", 2)
    repository.save_suite_result(
        run_id,
        0,
        {
            "case_id": "DEMO-01", "title": "Login", "module": "Đăng nhập",
            "page_key": "plt_login", "expected": "/login", "actual": "/login",
            "status": "PASSED", "message": "OK", "duration_ms": 100,
        },
    )
    repository.save_suite_result(
        run_id,
        1,
        {
            "case_id": "DEMO-02", "title": "Dashboard", "module": "Dashboard",
            "page_key": "plt_dashboard", "expected": "/dashboard", "actual": "",
            "status": "SKIPPED", "message": "No locator", "duration_ms": 0,
        },
    )
    summary = repository.finish_suite_run(run_id)

    assert summary["status"] == "PASSED_WITH_SKIPS"
    assert summary["passed"] == 1
    assert summary["skipped"] == 1
    assert len(repository.suite_run_results(run_id)) == 2
    assert len(repository.suite_run_module_summary(run_id)) == 2


def test_excel_and_html_reports_contain_summary_and_details():
    repository = _fresh_repository()
    suite_id = repository.save_suite_definition(
        "Report unit", demo_7_module_cases()[:1], suite_key="unit:report"
    )
    run_id = repository.start_suite_run(suite_id, "Report unit", "Full Website", 1)
    repository.save_suite_result(
        run_id,
        0,
        {
            "case_id": "DEMO-01", "title": "Login", "module": "Đăng nhập",
            "page_key": "plt_login", "expected": "/login", "actual": "/login",
            "status": "PASSED", "message": "OK", "log_text": "OPEN /login",
            "comparison_json": (
                '[{"expected":"/login","actual":"/login","status":"PASS"}]'
            ),
        },
    )
    run = repository.finish_suite_run(run_id)
    results = repository.suite_run_results(run_id)
    modules = repository.suite_run_module_summary(run_id)
    html_path = Path("data/test_suite_report.html")
    excel_path = Path("data/test_suite_report.xlsx")

    export_html(str(html_path), run, results, modules)
    export_excel(str(excel_path), run, results, modules)

    assert "TEST SUITE REPORT" in html_path.read_text(encoding="utf-8")
    assert "DEMO-01" in html_path.read_text(encoding="utf-8")
    assert "Expected Result" in html_path.read_text(encoding="utf-8")
    assert "Actual Result" in html_path.read_text(encoding="utf-8")
    assert excel_path.stat().st_size > 1000
    workbook = load_workbook(excel_path)
    assert "Comparison" in workbook.sheetnames
