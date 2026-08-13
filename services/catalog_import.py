"""Import test data cho module Danh mục xe của Linh.

Chuẩn hóa CSV / JSON / XLSX về:
    {"loai": "hang"|"mau", "ten": str, "hang": str, "trang_thai": str}
"""

from __future__ import annotations

import csv
import json
import unicodedata
from pathlib import Path


class CatalogImportError(ValueError):
    pass


def _plain(value: object) -> str:
    text = str(value or "").strip().lower()
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    return text.replace("đ", "d").replace(" ", "_")


COLUMN_ALIASES = {
    "loai": {"loai", "type", "kind"},
    "ten": {"ten", "name", "ten_xe", "tenxe"},
    "hang": {"hang", "brand", "hang_lien_ket", "hangxe"},
    "trang_thai": {"trang_thai", "status"},
}


def _normalize_type(value: object) -> str:
    key = _plain(value)
    if key in {"hang", "hang_xe", "brand"}:
        return "hang"
    if key in {"mau", "mau_xe", "model"}:
        return "mau"
    return key


def _map_row(raw: dict) -> dict:
    mapped = {}
    for raw_key, value in raw.items():
        key = _plain(raw_key)
        for target, aliases in COLUMN_ALIASES.items():
            if key == target or key in aliases:
                mapped[target] = value
                break

    return {
        "loai": _normalize_type(mapped.get("loai")),
        "ten": str(mapped.get("ten") or "").strip(),
        "hang": str(mapped.get("hang") or "").strip(),
        "trang_thai": str(mapped.get("trang_thai") or "Đang hoạt động").strip()
        or "Đang hoạt động",
    }


def _read_csv(path: Path) -> list[dict]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def _read_json(path: Path) -> list[dict]:
    with path.open("r", encoding="utf-8") as handle:
        payload = json.load(handle)
    if isinstance(payload, dict):
        payload = payload.get("rows", payload.get("data", []))
    if not isinstance(payload, list):
        raise CatalogImportError("JSON phải là danh sách hoặc có khóa rows/data.")
    return payload


def _read_xlsx(path: Path) -> list[dict]:
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        raise CatalogImportError(
            "Thiếu openpyxl. Cài bằng: pip install openpyxl"
        ) from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(value or "").strip() for value in rows[0]]
    data = []
    for values in rows[1:]:
        data.append({headers[i]: values[i] for i in range(min(len(headers), len(values)))})
    return data


def import_catalog_file(file_path: str | Path) -> tuple[list[dict], list[str]]:
    path = Path(file_path)
    if not path.exists():
        raise CatalogImportError(f"Không tìm thấy file: {path}")

    ext = path.suffix.lower()
    if ext == ".csv":
        raw_rows = _read_csv(path)
    elif ext == ".json":
        raw_rows = _read_json(path)
    elif ext == ".xlsx":
        raw_rows = _read_xlsx(path)
    else:
        raise CatalogImportError("Chỉ hỗ trợ CSV, JSON và XLSX.")

    rows: list[dict] = []
    warnings: list[str] = []
    for index, raw in enumerate(raw_rows, start=2):
        row = _map_row(raw)
        if row["loai"] not in {"hang", "mau"}:
            warnings.append(f"Dòng {index}: loại phải là Hãng/Mẫu xe.")
            continue
        if not row["ten"]:
            warnings.append(f"Dòng {index}: thiếu tên.")
            continue
        if row["loai"] == "mau" and not row["hang"]:
            warnings.append(f"Dòng {index}: mẫu xe '{row['ten']}' chưa có hãng liên kết.")
        rows.append(row)

    if not rows:
        raise CatalogImportError("Không có dữ liệu Danh mục xe hợp lệ.")
    return rows, warnings
